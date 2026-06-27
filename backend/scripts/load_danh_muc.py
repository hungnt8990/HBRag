"""Nạp danh mục tổ chức EVNCPC từ Excel vào PostgreSQL.

Đọc hai file:
- ``data/DM_DONVI.xlsx``    (sheet "Export Worksheet"): đơn vị x phòng ban.
- ``data/DM_NHANVIEN.xlsx`` (sheet đầu tiên): nhân viên.

Nạp vào ba bảng danh mục ``dm_don_vi`` / ``dm_phong_ban`` / ``dm_nhan_vien``
bằng upsert (ON CONFLICT DO UPDATE) nên chạy lại nhiều lần đều an toàn.

Chạy:
    python -m scripts.load_danh_muc
    python -m scripts.load_danh_muc --donvi data/DM_DONVI.xlsx --nhanvien data/DM_NHANVIEN.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import logging
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db.base import Base
from app.db.session import AsyncSessionLocal, engine
from app.models.danh_muc import DonVi, NhanVien, PhongBan

logger = logging.getLogger("load_danh_muc")

BATCH = 1000


def _int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    return {str(h).strip().lower(): i for i, h in enumerate(header_row) if h is not None}


def _build_org_paths(don_vi: dict[int, dict[str, Any]]) -> None:
    """Tính materialized path "/<root>/.../<id_dv>/" cho mỗi đơn vị."""
    for id_dv, row in don_vi.items():
        chain: list[int] = []
        seen: set[int] = set()
        cur: int | None = id_dv
        while cur is not None and cur not in seen and cur in don_vi:
            seen.add(cur)
            chain.append(cur)
            cur = don_vi[cur]["id_dv_cha"]
        chain.reverse()
        row["org_path"] = "/" + "/".join(str(c) for c in chain) + "/"


def parse_don_vi(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Trả về (danh sách đơn vị, danh sách phòng ban) từ DM_DONVI.xlsx."""
    # read_only=False vì workbook thiếu metadata dimension -> read_only đọc sai.
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb["Export Worksheet"] if "Export Worksheet" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    col = _header_map(next(rows))
    wb_close = wb.close

    def g(row: tuple[Any, ...], key: str) -> Any:
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    don_vi: dict[int, dict[str, Any]] = {}
    phong_ban: dict[int, dict[str, Any]] = {}
    for row in rows:
        if not row:
            continue
        id_dv = _int(g(row, "id_dv"))
        if id_dv is None:
            continue
        if id_dv not in don_vi:
            don_vi[id_dv] = {
                "id_dv": id_dv,
                "id_dv_cha": _int(g(row, "id_dv_cha")),
                "ma_dv": _str(g(row, "ma_dv")),
                "ky_hieu": _str(g(row, "ky_hieu")),
                "ten_dv": _str(g(row, "ten_dv")),
                "org_path": None,
            }
        id_pb = _int(g(row, "id_pb"))
        if id_pb is not None and id_pb not in phong_ban:
            phong_ban[id_pb] = {
                "id_pb": id_pb,
                "id_dv": id_dv,
                "ma_pb": _str(g(row, "ma_pb")),
                "ky_hieu_pb": _str(g(row, "ky_hieu_pb")),
                "ten_pb": _str(g(row, "ky_hieu_pb")),
            }
    wb_close()
    _build_org_paths(don_vi)
    return list(don_vi.values()), list(phong_ban.values())


def parse_nhan_vien(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    col = _header_map(next(rows))

    def g(row: tuple[Any, ...], key: str) -> Any:
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    nhan_vien: dict[int, dict[str, Any]] = {}
    for row in rows:
        if not row:
            continue
        id_nv = _int(g(row, "id_nv"))
        if id_nv is None:
            continue
        nhan_vien[id_nv] = {
            "id_nv": id_nv,
            "username": _str(g(row, "username")),
            "ho_ten": _str(g(row, "firstname")),
            "ten_hien_thi": _str(g(row, "lastname")),
            "email": _str(g(row, "email")),
            "id_dv": _int(g(row, "id_dv")),
            "ten_dv": _str(g(row, "ten_dv")),
            "id_pb": _int(g(row, "id_pb")),
            "ma_pb": _str(g(row, "ma_pb")),
            "ten_pb": _str(g(row, "ten_pb")),
            "id_hrms": _str(g(row, "id_hrms")),
            "ma_cv": _str(g(row, "ma_cv")),
            "ten_cv": _str(g(row, "ten_cv")),
            "is_active": True,
        }
    wb.close()
    return list(nhan_vien.values())


async def _upsert(session, model, rows: list[dict[str, Any]], pk: str) -> None:
    if not rows:
        return
    for start in range(0, len(rows), BATCH):
        batch = rows[start : start + BATCH]
        stmt = pg_insert(model).values(batch)
        update_cols = {c: stmt.excluded[c] for c in batch[0] if c != pk}
        stmt = stmt.on_conflict_do_update(index_elements=[pk], set_=update_cols)
        await session.execute(stmt)
    await session.commit()


async def ensure_tables() -> None:
    """Tạo 3 bảng danh mục nếu chưa có (checkfirst, không đụng alembic_version).

    Dùng khi chuỗi migration Alembic của branch đang lệch với DB. Khi đã reconcile
    được migration thì nên dùng `alembic upgrade head` thay cho bước này.
    """
    tables = [m.__table__ for m in (DonVi, PhongBan, NhanVien)]
    async with engine.begin() as conn:
        await conn.run_sync(lambda c: Base.metadata.create_all(c, tables=tables, checkfirst=True))
    logger.info("Đã đảm bảo 3 bảng danh mục tồn tại.")


async def load(donvi_path: Path, nhanvien_path: Path, *, create_tables: bool = True) -> None:
    if create_tables:
        await ensure_tables()
    logger.info("Đọc %s", donvi_path)
    don_vi, phong_ban = parse_don_vi(donvi_path)
    logger.info("  -> %d đơn vị, %d phòng ban", len(don_vi), len(phong_ban))

    logger.info("Đọc %s", nhanvien_path)
    nhan_vien = parse_nhan_vien(nhanvien_path)
    logger.info("  -> %d nhân viên", len(nhan_vien))

    async with AsyncSessionLocal() as session:
        await _upsert(session, DonVi, don_vi, "id_dv")
        await _upsert(session, PhongBan, phong_ban, "id_pb")
        await _upsert(session, NhanVien, nhan_vien, "id_nv")
    logger.info("Hoàn tất nạp danh mục.")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    base = Path(__file__).resolve().parent.parent / "data"
    parser = argparse.ArgumentParser(description="Nạp danh mục tổ chức EVNCPC từ Excel vào PostgreSQL.")
    parser.add_argument("--donvi", type=Path, default=base / "DM_DONVI.xlsx")
    parser.add_argument("--nhanvien", type=Path, default=base / "DM_NHANVIEN.xlsx")
    parser.add_argument(
        "--no-create",
        action="store_true",
        help="Bỏ qua bước tạo bảng (dùng khi bảng đã được tạo qua Alembic).",
    )
    args = parser.parse_args()
    asyncio.run(load(args.donvi, args.nhanvien, create_tables=not args.no_create))


if __name__ == "__main__":
    main()
